from exceptions import NoPhoneReportFound
from net_async import multithread
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from datetime import datetime
import re


class CdpParser:
    """
    Parses outputs of commands: 'show cdp neighbor', 'show interface switchport', and 'show mac address-table'.
        Attributes:
            phones = []\n
            routers_switches = []\n
            waps = []\n
            others = []\n
        Dictionary format within lists:
            {
                'hostname',\n
                'ip_address',\n
                'model',\n
                'software_version',\n
                'neighbor': { (on router_switch, intfs are not in 'neighbor')
                    'hostname',\n
                    'ip_address',\n
                    'remote_intf', (neighbor interface)\n
                    'local_intf', (local to device, not on wap or phone)\n
                }\n
                'mac_addr', (phone only)\n
                'voice_vlan', (phone only)
            }
    """

    def __init__(self, cdp_neighbors, switchports, mac_addrs, session):
        nxos = False
        try:
            _ = cdp_neighbors[0]['destination_host']
            hostname_s = 'destination_host'
            version_s = 'software_version'
            mgmt_ip_s = 'management_ip'
        except KeyError:
            nxos = True
            hostname_s = 'dest_host'
            version_s = 'version'
            mgmt_ip_s = 'mgmt_ip'
        self.phones = []
        self.routers_switches = []
        self.waps = []
        self.others = []

        def phone_parse(neighbor):
            """Returns dictionary for CDP neighbor phone"""
            mgmt_ip = neighbor[mgmt_ip_s]
            hostname = neighbor[hostname_s].split('.')[0]
            if nxos:
                sysname = neighbor['sysname']
                if sysname != '':
                    hostname = sysname
                if mgmt_ip == '':
                    mgmt_ip = neighbor['interface_ip']
            l_intf = neighbor['local_port']
            intf = re.findall(r'.{2}', l_intf)[0] + re.findall(r'\d.+', l_intf)[0]
            macreg = re.findall(r'.{4}', hostname.replace('SEP', ''))
            mac_address = f'{macreg[0]}.{macreg[1]}.{macreg[2]}'.lower()
            voice_vlan = 'None'
            software_version = neighbor[version_s].replace('.loads', '')
            platform = neighbor['platform']
            for switchport in switchports:
                if switchport['interface'] == intf:
                    for mac_addr in mac_addrs:
                        if mac_addr['vlan'] == switchport['voice_vlan']:
                            voice_vlan = mac_addr['vlan']
                            break
                    break
            if platform.__contains__('Cisco IP Phone'):
                platform = neighbor['platform'].replace('Cisco IP Phone ', '')
            else:
                platform = neighbor['platform']
            phone = {
                'hostname': hostname,
                'neighbor': {
                    'hostname': session.hostname,
                    'ip_address': session.ip_address,
                    'remote_intf': l_intf
                },
                'ip_address': mgmt_ip,
                'mac_addr': mac_address,
                'voice_vlan': voice_vlan,
                'software_version': software_version,
                'model': platform
            }
            self.phones.append(phone)

        def router_sw_parse(neighbor):
            """Returns dictionary for CDP neighbor router or switch"""
            mgmt_ip = neighbor[mgmt_ip_s]
            hostname = neighbor[hostname_s].split('.')[0]
            if hostname.__contains__('('):
                hostname = hostname.split('(')[0]
            if nxos:
                sysname = neighbor['sysname']
                if sysname != '':
                    hostname = sysname
                if mgmt_ip == '':
                    mgmt_ip = neighbor['interface_ip']
            software_version = neighbor[version_s]
            platform = neighbor['platform']
            for software in software_version.split(','):
                if software.__contains__('Version'):
                    software_version = software.split('Version')[1].split('REL')[0]
                    if software_version.__contains__(':'):
                        software_version = software_version.replace(': ', '')
                    else:
                        software_version = software_version.replace(' ', '')
                    break
            if platform.__contains__('cisco '):
                platform = neighbor['platform'].replace('cisco ', '')
            elif platform.__contains__('Cisco '):
                platform = neighbor['platform'].replace('Cisco ', '')
            else:
                platform = neighbor['platform']
            router_sw = {
                'hostname': hostname,
                'ip_address': mgmt_ip,
                'remote_intf': neighbor['local_port'],
                'local_intf': neighbor['remote_port'],
                'software_version': software_version,
                'model': platform
            }
            self.routers_switches.append(router_sw)

        def wap_parse(neighbor):
            """Returns dictionary for CDP neighbor wireless access point"""
            mgmt_ip = neighbor[mgmt_ip_s]
            hostname = neighbor[hostname_s].split('.')[0]
            if nxos:
                sysname = neighbor['sysname']
                if sysname != '':
                    hostname = sysname
                if mgmt_ip == '':
                    mgmt_ip = neighbor['interface_ip']
            software_version = neighbor[version_s]
            platform = neighbor['platform']
            for software in software_version.split(','):
                if software.__contains__('Version'):
                    software_version = software.split('Version')[1]
                    if software_version.__contains__(':'):
                        software_version = software_version.replace(': ', '')
                    else:
                        software_version = software_version.replace(' ', '')
                    break
            if platform.__contains__('cisco '):
                platform = neighbor['platform'].replace('cisco ', '')
            elif platform.__contains__('Cisco '):
                platform = neighbor['platform'].replace('Cisco ', '')
            else:
                platform = neighbor['platform']
            ap = {
                'hostname': hostname,
                'ip_address': mgmt_ip,
                'model': platform,
                'neighbor': {
                    'hostname': session.hostname,
                    'ip_address': session.ip_address,
                    'remote_intf': neighbor['local_port']
                },
                'software_version': software_version
            }
            self.waps.append(ap)

        def other_parse(neighbor):
            """Returns dictionary for CDP neighbor that isn't a phone, access point, router, or switch"""
            mgmt_ip = neighbor[mgmt_ip_s]
            hostname = neighbor[hostname_s].split('.')[0]
            if nxos:
                sysname = neighbor['sysname']
                if sysname != '':
                    hostname = sysname
                if mgmt_ip == '':
                    mgmt_ip = neighbor['interface_ip']
            software_version = neighbor[version_s]
            if software_version.__contains__(','):
                for software in software_version.split(','):
                    if software.__contains__('Version'):
                        software_version = software.split('Version')[1].split('REL')[0]
                        if software_version.__contains__(':'):
                            software_version = software_version.replace(': ', '')
                        else:
                            software_version = software_version.replace(' ', '')
                        break
            elif software_version.__contains__('Version'):
                found_1 = False
                for x in software_version.split(' '):
                    if x.__contains__('Version'):
                        found_1 = True
                        continue
                    if found_1:
                        software_version = x
                        break
            elif software_version.__contains__('version'):
                found_1 = False
                for x in software_version.split(' '):
                    if x.__contains__('version'):
                        found_1 = True
                        continue
                    if found_1:
                        software_version = x
                        break
            platform = neighbor['platform']
            if platform.__contains__('cisco '):
                platform = neighbor['platform'].replace('cisco ', '')
            elif platform.__contains__('Cisco '):
                platform = neighbor['platform'].replace('Cisco ', '')
            else:
                platform = neighbor['platform']
            other = {
                'hostname': hostname,
                'ip_address': mgmt_ip,
                'neighbor': {
                    'hostname': session.hostname,
                    'ip_address': session.ip_address,
                    'remote_intf': neighbor['local_port'],
                    'local_intf': neighbor['remote_port']
                },
                'software_version': software_version,
                'model': platform
            }
            self.others.append(other)

        def parse(n):
            """Given TEXTFSM CDP neighbor, checks type of device and runs through corresponding parser function."""
            capabilities = n['capabilities']
            if n['platform'].__contains__('IP Phone') or capabilities.__contains__('Phone'):
                phone_parse(n)
            elif capabilities.__contains__('Router') and capabilities.__contains__('Source-Route-Bridge') or \
                    capabilities.__contains__('Switch'):
                router_sw_parse(n)
            elif capabilities.__contains__('Trans-Bridge'):
                wap_parse(n)
            else:
                other_parse(n)

        multithread(parse, cdp_neighbors)


def cucm_export_parse(file):
    """Parses CUCM export of phones with fields 'Description', 'Device Name', and 'Directory Number 1'

    :returns:  {'SEP000000000000': {'description', 'directory_number'}}"""
    phones = {}
    while True:
        try:
            with open(file) as phonelist_csv:
                for line in phonelist_csv:
                    if not line.__contains__('Description,Device Name,Directory Number 1'):
                        info = line.split(',')
                        device_name = info[1]
                        description = info[0]
                        directory_number = info[2]
                        phones[device_name.upper()] = {
                            'description': description,
                            'directory_number': directory_number
                        }
            return phones
        except FileNotFoundError:
            raise NoPhoneReportFound('No phone report file found at provided location.')


def output_to_spreadsheet(routers_switches, phones, aps, others, failed_devices, file_location):
    """Parses device lists and outputs to spreadsheet"""
    # Creates Excel workbook and worksheets
    wb = Workbook()
    routers_switches_ws = wb.active
    routers_switches_ws.title = 'Routers_Switches'
    phones_ws = wb.create_sheet('Phones')
    aps_ws = wb.create_sheet('APs')
    others_ws = wb.create_sheet('Others')
    failed_ws = wb.create_sheet('Failed')

    alphabet = 'AAABACADAEAFAGAHAIAJAKALAMANAOAPAQARASATAUAVAWAXAYAZBABBBCBDBEBFBGBHBIBJBKBLBMBNBOBPBQBRBSBTBUBVBWBXB' \
               'YBZCACBCCCDCECFCGCHCICJCKCLCMCNCOCPCQCRCSCTCUCVCWCXCYCZDADBDCDDDEDFDGDHDIDJDKDLDMDNDODPDQDRDSDTDUDVDW' \
               'DXDYDZEAEBECEDEEEFEGEHEIEJEKELEMENEOEPEQERESETEUEVEWEXEYEZFAFBFCFDFEFFFGFHFIFJFKFLFMFNFOFPFQFRFSFTFUF' \
               'VFWFXFYFZGAGBGCGDGEGFGGGHGIGJGKGLGMGNGOGPGQGRGSGTGUGVGWGXGYGZHAHBHCHDHEHFHGHHHIHJHKHLHMHNHOHPHQHRHSHT' \
               'HUHVHWHXHYHZIAIBICIDIEIFIGIHIIIJIKILIMINIOIPIQIRISITIUIVIWIXIYIZJAJBJCJDJEJFJGJHJIJJJKJLJMJNJOJPJQJRJ' \
               'SJTJUJVJWJXJYJZKAKBKCKDKEKFKGKHKIKJKKKLKMKNKOKPKQKRKSKTKUKVKWKXKYKZLALBLCLDLELFLGLHLILJLKLLLMLNLOLPLQ' \
               'LRLSLTLULVLWLXLYLZMAMBMCMDMEMFMGMHMIMJMKMLMMMNMOMPMQMRMSMTMUMVMWMXMYMZNANBNCNDNENFNGNHNINJNKNLNMNNNON' \
               'PNQNRNSNTNUNVNWNXNYNZOAOBOCODOEOFOGOHOIOJOKOLOMONOOOPOQOROSOTOUOVOWOXOYOZPAPBPCPDPEPFPGPHPIPJPKPLPMPN' \
               'POPPPQPRPSPTPUPVPWPXPYPZQAQBQCQDQEQFQGQHQIQJQKQLQMQNQOQPQQQRQSQTQUQVQWQXQYQZRARBRCRDRERFRGRHRIRJRKRLR' \
               'MRNRORPRQRRRSRTRURVRWRXRYRZSASBSCSDSESFSGSHSISJSKSLSMSNSOSPSQSRSSSTSUSVSWSXSYSZTATBTCTDTETFTGTHTITJTK' \
               'TLTMTNTOTPTQTRTSTTTUTVTWTXTYTZUAUBUCUDUEUFUGUHUIUJUKULUMUNUOUPUQURUSUTUUUVUWUXUYUZVAVBVCVDVEVFVGVHVIV' \
               'JVKVLVMVNVOVPVQVRVSVTVUVVVWVXVYVZWAWBWCWDWEWFWGWHWIWJWKWLWMWNWOWPWQWRWSWTWUWVWWWXWYWZXAXBXCXDXEXFXGXH' \
               'XIXJXKXLXMXNXOXPXQXRXSXTXUXVXWXXXYXZYAYBYCYDYEYFYGYHYIYJYKYLYMYNYOYPYQYRYSYTYUYVYWYXYYYZZAZBZCZDZEZFZ' \
               'GZHZIZJZKZLZMZNZOZPZQZRZSZTZUZVZWZXZYZZ'

    # Checks if phones contain directory number and description from CUCM export merge
    if any('description' in phone for phone in phones):
        phone_string = 'CUCMPhone'
    else:
        phone_string = 'Phone'

    neighbor_count = 1
    # Sets 'neighbor_count' to length of longest neighbor list in routers_switches dictionaries
    for rt_sw in routers_switches:
        if rt_sw['connection_attempt'] == 'Failed':
            if len(rt_sw['neighbors']) > neighbor_count:
                neighbor_count = len(rt_sw['neighbors'])

    def write_header(worksheet, device_type):
        """
        :param device_type: 'RouterSwitch', 'Phone', 'CUCMPhone', 'WAP', 'Other', or 'Failed'
        :param worksheet: Device worksheet
        :return: int(header_length), list(header)
        """
        header = ['Hostname', 'IP Address', 'Model', 'Software Version']
        if device_type == 'RouterSwitch':
            header += ['Serial', 'Connection Type', 'ROMMON', 'Connection Attempt', 'Discovery Status']
            for n in range(1, neighbor_count + 1):
                header += [f'Neighbor {n} Hostname', f'Neighbor {n} IP Address', f'Local Interface to Neighbor {n}',
                           f'Neighbor {n} Interface']
        elif device_type == 'Phone' or device_type == 'CUCMPhone':
            header += ['Voice VLAN', 'MAC Address', 'Switch Hostname', 'Switch IP Address', 'Switchport']
            if device_type == 'CUCMPhone':
                header += ['Description', 'Main Directory Number']
        elif device_type == 'WAP':
            header += ['Switch Hostname', 'Switch IP Address', 'Switchport']
        elif device_type == 'Other':
            header += ['Neighbor  Hostname', 'Neighbor IP Address', 'Local Interface to Neighbor', 'Neighbor Interface']
        elif device_type == 'Failed':
            header = ['IP Address', 'Connection Type', 'Device Type', 'Connectivity', 'Authentication',
                      'Authorization', 'Discovery Status', 'Connection Exception']
        worksheet.append(header)
        return len(header), header

    def write_to_sheet(device_list, worksheet, device_type):
        """
        :param device_type: 'RouterSwitch', 'Phone', 'CUCMPhone', 'WAP', 'Other', or 'Failed'
        :param device_list: List of devices
        :param worksheet: Device worksheet
        :return: list(rows)
        """
        rows = []
        for device in device_list:
            if device_type != 'Failed':
                row = [device['hostname'], device['ip_address'], device['model'], device['software_version']]
                if device_type == 'RouterSwitch':
                    if 'serial' in device:
                        serial = device['serial']
                        connection_type = device['connection_type']
                        rommon = device['rommon']
                    else:
                        serial = 'Unknown'
                        connection_type = 'Unknown'
                        rommon = 'Unknown'
                    row += [serial, connection_type, rommon, device['connection_attempt'], device['discovery_status']]
                    if device['connection_attempt'] == 'Failed':
                        for neighbor in device['neighbors']:
                            row += [neighbor['hostname'], neighbor['ip_address'], neighbor['local_intf'],
                                    neighbor['remote_intf']]
                if device_type == 'Phone' or device_type == 'CUCMPhone':
                    neighbor = device['neighbor']
                    row += [device['voice_vlan'], device['mac_addr'], neighbor['hostname'], neighbor['ip_address'],
                            neighbor['remote_intf']]
                    if 'description' in device:
                        row += [device['description'], device['directory_number']]
                if device_type == 'WAP' or device_type == 'Other':
                    neighbor = device['neighbor']
                    row += [neighbor['hostname'], neighbor['ip_address'], neighbor['remote_intf']]
                    if device_type == 'Other':
                        row.append(neighbor['local_intf'])
            else:
                row = [device['ip_address'], device['connection_type'], device['device_type'], device['connectivity'],
                       device['authentication'], device['authorization'], device['discovery_status'],
                       device['exception']]
            worksheet.append(row)
            rows.append(row)
        return rows

    def complete_sheet(device_list, worksheet, device_type):
        """Completes workbook sheet"""
        column_num = len(device_list) + 1
        header_out = write_header(worksheet, device_type)
        header = header_out[1]
        header_length = header_out[0]
        letter = header_length - 1
        column_letter = alphabet[letter]
        bottom_right_cell = f'{column_letter}{column_num}'
        rows = write_to_sheet(device_list, worksheet, device_type)

        # Creates table if there is data in table
        if len(device_list) != 0:
            table = Table(displayName=device_type, ref=f'A1:{bottom_right_cell}')
            style = TableStyleInfo(name='TableStyleMedium9', showFirstColumn=False, showLastColumn=False,
                                   showRowStripes=True, showColumnStripes=True)
            table.tableStyleInfo = style
            worksheet.add_table(table)

        # Sets column widths
        all_data = [header]
        all_data += rows
        column_widths = []
        for row in all_data:
            for i, cell in enumerate(row):
                if len(column_widths) > i:
                    if len(str(cell)) > column_widths[i]:
                        column_widths[i] = len(str(cell))
                else:
                    column_widths += [len(str(cell))]

        for i, column_width in enumerate(column_widths):
            worksheet.column_dimensions[alphabet[i]].width = column_width + 3

    complete_sheet(routers_switches, routers_switches_ws, 'RouterSwitch')
    complete_sheet(phones, phones_ws, phone_string)
    complete_sheet(aps, aps_ws, 'WAP')
    complete_sheet(others, others_ws, 'Other')
    complete_sheet(failed_devices, failed_ws, 'Failed')

    # Saves workbook
    date_time = datetime.now().strftime('%m_%d_%Y-%H_%M_%S')
    wb.save(f'{file_location}/network_inventory-{date_time}-.xlsx')
